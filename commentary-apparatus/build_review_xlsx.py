#!/usr/bin/env python3
"""Build 'Reconcile decisions.xlsx' — the human review surface over the gate.
Generated from reconciliation.json + works_cited_staged.json. Filling the DECISION
dropdowns and running ingest_decisions.py reproduces reconcile_decisions.json.
"""
import json, os
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

AP = os.environ.get("APPARATUS_DIR", os.getcwd())
recon = json.load(open(os.path.join(AP, "reconciliation.json")))
staged = json.load(open(os.path.join(AP, "works_cited_staged.json")))["works"]
res = recon["id_resolution"]

HELD = {"levin-2026-ferriss-interview", "carroll-2026-mindscape-349-harlow"}
FNT = "Arial"
hdr_fill = PatternFill("solid", fgColor="1F3864"); hdr_font = Font(name=FNT, bold=True, color="FFFFFF", size=11)
input_fill = PatternFill("solid", fgColor="FFF2CC")   # yellow = you edit
hold_fill  = PatternFill("solid", fgColor="F8CBAD")   # orange = decide before finishing
ref_fill   = PatternFill("solid", fgColor="E2EFDA")
thin = Side(style="thin", color="BFBFBF"); border = Border(thin, thin, thin, thin)
wrap = Alignment(wrap_text=True, vertical="top")

wb = Workbook()

def style_header(ws, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(1, c); cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

# ---------------- Tab 1: How to ----------------
ws = wb.active; ws.title = "How to"
lines = [
 ("Reconcile decisions — how to fill this in", 14, True),
 ("", 11, False),
 ("This is the human surface over the step-3 reconcile gate. Fill the yellow DECISION cells on the 'Decide' tab, then finish (below).", 11, False),
 ("Everything else — 'PRS resolutions', 'CROSS-FLAG', — is read-only reference so you can spot exceptions.", 11, False),
 ("", 11, False),
 ("LEGEND", 12, True),
 ("  Yellow cell  = you edit it (a dropdown or a value).", 11, False),
 ("  Orange row   = a genuine judgment call, pre-set to 'hold'. The gate will NOT close while any row is 'hold'.", 11, False),
 ("  Green tab    = read-only reference, no decision needed.", 11, False),
 ("", 11, False),
 ("THE DECIDE TAB (12 rows)", 12, True),
 ("  6 staged works — DECISION: approve | amend | decline | hold.", 11, False),
 ("     amend  -> put corrected text in 'Value / redirect' (e.g. a fixed year or author).", 11, False),
 ("     decline-> put in 'Value / redirect' the cite_key the dependent PRS should point to instead.", 11, False),
 ("  6 batch policies — accept the recommendation or change the one word in the dropdown.", 11, False),
 ("", 11, False),
 ("EXCEPTIONS", 12, True),
 ("  To override a single PRS's target, add a row on the 'Overrides' tab: the PRS id + the cite_key you want.", 11, False),
 ("", 11, False),
 ("WHEN FINISHED (pick one)", 12, True),
 ("  A. Tell me 'decisions done' in chat — I stage this sheet, apply it, and hand back the resolved files + a git diff. (lowest friction)", 11, False),
 ("  B. Run locally:  python3 ingest_decisions.py 'Reconcile decisions.xlsx'  — writes reconcile_decisions.json and runs the merge.", 11, False),
 ("  C. Want true hands-off? Say so and I'll host this as a Google Sheet and pull it on a schedule when you flip a Submit cell.", 11, False),
 ("", 11, False),
 ("The merge is non-destructive and fail-loud: it never rewrites works_cited.json, and it refuses to report 'done' while anything is held.", 11, False),
]
for i,(t,sz,b) in enumerate(lines, 1):
    c = ws.cell(i,1,t); c.font = Font(name=FNT, size=sz, bold=b)
ws.column_dimensions["A"].width = 130

# ---------------- Tab 2: Decide ----------------
ws = wb.create_sheet("Decide")
cols = ["Item", "Kind", "What it is  /  recommendation", "DECISION", "Value / redirect", "Notes"]
ws.append(cols)
widths = [42, 12, 60, 16, 30, 44]
for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width = w
style_header(ws, len(cols))

dv_work = DataValidation(type="list", formula1='"approve,amend,decline,hold"', allow_blank=False)
ws.add_data_validation(dv_work)
policy_dvs = {}

r = 2
# staged works
for k, w in staged.items():
    rec = f'{", ".join(w["authors"])}, "{w["title"]}" ({w["work_type"]}, {w["year"]}). REC: ' + ("approve" if k not in HELD else "HOLD — judgment call")
    dec = "hold" if k in HELD else "approve"
    row = [k, "staged work", rec + " — " + (w.get("note","")), dec, "", ""]
    ws.append(row)
    dv_work.add(ws.cell(r,4))
    ws.cell(r,4).fill = input_fill; ws.cell(r,5).fill = input_fill
    if k in HELD:
        for c in range(1,7): ws.cell(r,c).fill = hold_fill
        ws.cell(r,4).fill = input_fill
    for c in range(1,7): ws.cell(r,c).border=border; ws.cell(r,c).alignment=wrap
    r += 1

# batch policies
POL = [
 ("rc_tome_prs_to_canonical","accept","accept,review_each","24 RC-Tome PRS re-descriptions -> thinker canonical work. REC: accept"),
 ("existing_seeded_prs","accept","accept,review_each","2 PRS ids map to an already-seeded work (hoffman-07, wolfram-06). REC: accept"),
 ("generics_canonical_default","accept","accept,review_each","484 generic surname mentions stay at thinker canonical (foundation §4). REC: accept"),
 ("friston_active_inference","keep_generic","keep_generic,promote_all,per_day","44 Friston days matched the denylisted concept-label 'active inference'. REC: keep_generic (->FEP)"),
 ("unscoped_prs_scope","master_framework","master_framework,per_day","10 unscoped PRS ids (PRS-01..11). REC: master_framework (C2A2/Loughran form)"),
 ("cross_flag_internal","accept","accept,review_each","33 CROSS + 2 FLAG internal bridges, stripped at export §5. REC: accept (spot-check only)"),
]
for name, default, choices, desc in POL:
    ws.append([name, "policy", desc, default, "", ""])
    dv = DataValidation(type="list", formula1=f'"{choices}"', allow_blank=False)
    ws.add_data_validation(dv); dv.add(ws.cell(r,4))
    ws.cell(r,4).fill = input_fill
    for c in range(1,7): ws.cell(r,c).border=border; ws.cell(r,c).alignment=wrap
    r += 1

# ---------------- Tab 3: Overrides ----------------
ws = wb.create_sheet("Overrides")
ws.append(["PRS id (exception)", "New underlying_work cite_key", "Note"])
ws.append(["levin-PRS-01", "levin-2018-bioelectric-code", "EXAMPLE row — delete or overwrite. Prefer the bioelectric paper over canonical."])
for i,w in enumerate([26,40,60],1): ws.column_dimensions[chr(64+i)].width=w
style_header(ws,3)
ws.cell(2,1).font=Font(name=FNT, italic=True, color="808080")
ws.cell(2,2).font=Font(name=FNT, italic=True, color="808080")
ws.cell(2,3).font=Font(name=FNT, italic=True, color="808080")

# ---------------- Tab 4: PRS resolutions (ref) ----------------
ws = wb.create_sheet("PRS resolutions (ref)")
ws.append(["PRS id", "-> underlying work", "resolution", "days", "note"])
for i,w in enumerate([20,46,26,8,60],1): ws.column_dimensions[chr(64+i)].width=w
for pid, rr in sorted(res["scoped_prs"].items()):
    ws.append([pid, rr["underlying_work"], rr["resolution"], len(rr["days"]), rr["note"]])
style_header(ws,5)
for row in ws.iter_rows(min_row=2):
    for c in row: c.fill=ref_fill; c.alignment=wrap; c.font=Font(name=FNT,size=10)

# ---------------- Tab 5: CROSS-FLAG (ref) ----------------
ws = wb.create_sheet("CROSS-FLAG (ref)")
ws.append(["id", "programs", "-> canonical works if cited", "occ", "note"])
for i,w in enumerate([14,44,50,6,50],1): ws.column_dimensions[chr(64+i)].width=w
for uid, rr in sorted(res["cross"].items()):
    ws.append([uid, rr["programs"], ", ".join(rr["underlying_canonical_works"]), rr["n_occurrences"], "internal bridge (stripped at export)"])
for uid, rr in sorted(res["flag"].items()):
    ws.append([uid, "(paradigm flag)", "", rr["n_occurrences"], rr["note"]])
style_header(ws,5)
for row in ws.iter_rows(min_row=2):
    for c in row: c.fill=ref_fill; c.alignment=wrap; c.font=Font(name=FNT,size=10)

for name in ["Decide","PRS resolutions (ref)","CROSS-FLAG (ref)","Overrides"]:
    for row in wb[name].iter_rows(min_row=2):
        for c in row:
            if not c.font or c.font.name!=FNT: pass
out = os.path.join(AP, "Reconcile decisions.xlsx")
wb.save(out)
print("wrote", out)
