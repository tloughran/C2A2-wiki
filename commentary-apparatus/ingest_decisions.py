#!/usr/bin/env python3
"""
ingest_decisions.py — read 'Reconcile decisions.xlsx' -> reconcile_decisions.json -> run merge.

The spreadsheet is the human surface; this turns it back into the machine decision file the
gate already understands, then invokes merge_decisions.py. Non-destructive; fail-loud is
inherited from merge (nonzero exit while anything is 'hold').

Usage:  python3 ingest_decisions.py ["Reconcile decisions.xlsx"]
Paths via $APPARATUS_DIR (default cwd).
"""
import json, os, sys, subprocess
from openpyxl import load_workbook

AP = os.environ.get("APPARATUS_DIR", os.getcwd())
xlsx = sys.argv[1] if len(sys.argv) > 1 else os.path.join(AP, "Reconcile decisions.xlsx")
wb = load_workbook(xlsx)

# seed defaults/choices/notes from a fresh --init so we preserve schema + choices
subprocess.run([sys.executable, os.path.join(AP, "merge_decisions.py"), "--init"],
               cwd=AP, check=True, stdout=subprocess.DEVNULL)
dec = json.load(open(os.path.join(AP, "reconcile_decisions.json")))

# --- Decide tab: rows are staged works then policies, keyed by column A
ws = wb["Decide"]
staged_keys = set(dec["staged_works"]); policy_keys = set(dec["batch_policies"])
for r in range(2, ws.max_row + 1):
    item = (ws.cell(r,1).value or "").strip()
    decision = (ws.cell(r,4).value or "").strip()
    value = (ws.cell(r,5).value or "").strip()
    if item in staged_keys:
        d = dec["staged_works"][item]; d["decision"] = decision
        if decision == "decline": d["redirect"] = value or None
        elif decision == "amend" and value:
            # value formatted as "field=val; field2=val2"
            d["fields"] = {kv.split("=",1)[0].strip(): kv.split("=",1)[1].strip()
                           for kv in value.split(";") if "=" in kv}
    elif item in policy_keys:
        dec["batch_policies"][item]["decision"] = decision

# --- Overrides tab
ovr = {}
if "Overrides" in wb.sheetnames:
    o = wb["Overrides"]
    for r in range(2, o.max_row + 1):
        pid = (o.cell(r,1).value or "").strip()
        tgt = (o.cell(r,2).value or "").strip()
        note = (o.cell(r,3).value or "").strip()
        if pid and tgt and "EXAMPLE" not in note.upper():
            ovr[pid] = {"underlying_work": tgt, "note": note}
dec["overrides"] = ovr or {"_examples": {}}

json.dump(dec, open(os.path.join(AP, "reconcile_decisions.json"), "w"), indent=1, ensure_ascii=False)
n_hold = sum(1 for d in dec["staged_works"].values() if d["decision"] == "hold")
print(f"ingested {xlsx} -> reconcile_decisions.json  ({len(dec['staged_works'])} works, {n_hold} held, {len(ovr)} overrides)\n")
sys.exit(subprocess.run([sys.executable, os.path.join(AP,"merge_decisions.py")], cwd=AP).returncode)
